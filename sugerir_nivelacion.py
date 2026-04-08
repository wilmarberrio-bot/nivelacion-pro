import openpyxl
import os
import math
from datetime import datetime

from config import *
from normalization import *
from status_rules import *
from geo_utils import *
from routing_rules import *
from excel_export import (
    style_header_row,
    auto_fit_columns,
    ALERT_FILL,
    SUCCESS_FILL,
    WARN_FILL,
    Alignment,
)

print(">>> CARGANDO sugerir_nivelacion DESDE:", __file__)

def generate_suggestions(input_file, forced_hour=None):
    _now = now_bogota()
    if forced_hour is not None:
        current_hour = forced_hour
    else:
        current_hour = _now.hour + _now.minute / 60.0

    print(f"Leyendo archivo: {input_file}")
    print(f"Hora actual (COL): {current_hour:.2f} ({_now.strftime('%H:%M')} America/Bogota)")

    try:
        wb = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
        sheet = wb.active
    except Exception as e:
        return f"Error leyendo el archivo: {str(e)}", None

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    idx_status = get_col_index(headers, ['status_txt', 'Estado', 'Estado de orden de trabajo'])
    idx_zone = get_col_index(headers, ['Cities__name', 'Ciudad', 'Zona'])
    idx_zone_fallback = get_col_index(headers, ['Zone Name'])
    idx_zone_op = get_col_index(headers, ['zona_op'])
    idx_subzone = get_col_index(headers, ['Subzone', 'Subzona', 'Sites'])
    idx_tech = get_col_index(headers, ['tecnico', 'technenvician', 'Tecnico asignado', 'Tecnico', 'Tech'])
    idx_order = get_col_index(headers, ['appointment_id', 'ID', 'Numero de orden', 'Order ID'])
    idx_franja = get_col_index(headers, ['franja_label', 'Franja', 'Cita', 'Ventana'])
    idx_lat = get_col_index(headers, ['Latitude', 'Latitud', 'lat'])
    idx_lon = get_col_index(headers, ['Longitude', 'Longitud', 'lon', 'lng'])

    # ✅ NUEVO: link de Google Maps
    idx_maps = get_col_index(headers, ['Google Maps Link', 'google maps link', 'maps link', 'map link', 'GoogleMapsLink'])

    idx_address = get_col_index(headers, ['Addresses__address', 'Direccion', 'direccion'])
    idx_onsite = get_col_index(headers, ['onsite_at_cot'])

    missing = []
    if idx_tech == -1: missing.append("Tecnico")
    if idx_order == -1: missing.append("ID Orden")
    if missing:
        return f"Error: No se encontraron las columnas: {', '.join(missing)}", None

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tech = norm_text(row[idx_tech], "SIN_ASIGNAR") if (idx_tech != -1) else "SIN_ASIGNAR"
        if not tech or tech.lower() in ['none', 'nan', '']:
            tech = "SIN_ASIGNAR"

        # Zona: prioridad Cities__name > Zone Name > zona_op
        zona = "SIN_ZONA"
        if idx_zone != -1 and row[idx_zone] and str(row[idx_zone]).strip().lower() not in ['none', '']:
            zona = str(row[idx_zone]).strip()
        elif idx_zone_fallback != -1 and row[idx_zone_fallback] and str(row[idx_zone_fallback]).strip().lower() not in ['none', '']:
            zona = str(row[idx_zone_fallback]).strip()
        elif idx_zone_op != -1 and row[idx_zone_op] and str(row[idx_zone_op]).strip().lower() not in ['none', '']:
            zona = "Zona " + str(row[idx_zone_op]).strip()

        subzona = "SIN_SUBZONA"
        if idx_subzone != -1 and row[idx_subzone] and str(row[idx_subzone]).strip().lower() not in ['none', '']:
            subzona = str(row[idx_subzone]).strip()

        zona = norm_zone(zona)
        subzona = norm_subzone(subzona)

        order_id = row[idx_order] if idx_order != -1 else "N/A"
        status = norm_text(row[idx_status], "Sin Estado") if (idx_status != -1) else "Sin Estado"
        franja = norm_text(row[idx_franja], "Sin Franja") if (idx_franja != -1) else "Sin Franja"
        franja = franja.replace('\ufffd', '-').replace('\u2013', '-').replace('\u2014', '-')

        lat, lon = 0.0, 0.0
        if idx_lat != -1 and row[idx_lat]:
            try: lat = float(row[idx_lat])
            except: pass
        if idx_lon != -1 and row[idx_lon]:
            try: lon = float(row[idx_lon])
            except: pass

        address = norm_text(row[idx_address], "") if (idx_address != -1) else ""
        maps_link = norm_text(row[idx_maps], "") if (idx_maps != -1) else ""

        # ✅ FALLBACK: si no hay coords, intentar extraerlas del link de Google Maps y luego del address
        if (lat == 0.0 or lon == 0.0):
            lat2, lon2 = extract_coords_from_text(maps_link)
            if lat2 is not None and lon2 is not None:
                lat, lon = lat2, lon2
            else:
                lat3, lon3 = extract_coords_from_text(address)
                if lat3 is not None and lon3 is not None:
                    lat, lon = lat3, lon3

        addr_key = build_address_key(address, subzona)

        onsite_dt = row[idx_onsite] if idx_onsite != -1 else None
        onsite_hour = None
        if onsite_dt and isinstance(onsite_dt, datetime):
            onsite_hour = onsite_dt.hour + onsite_dt.minute / 60.0

        data.append({
            'tech': tech, 'zona': zona, 'subzona': subzona,
            'order_id': order_id, 'status': status, 'franja': franja,
            'lat': lat, 'lon': lon, 'address': address, 'maps_link': maps_link, 'addr_key': addr_key,
            'maps_present': bool(maps_link),
            'coords_ok': (lat != 0.0 and lon != 0.0),
            'sector': coords_to_sector(lat, lon, subzona),
            'onsite_hour': onsite_hour
        })

    wb.close()

    if not data:
        return "Error: No se encontraron datos en el archivo.", None

    print(f"Total ordenes: {len(data)}")

    zones = sorted(set(d['zona'] for d in data if d['zona'] != "SIN_ZONA"))

    unique_franjas = sorted(list(set(d['franja'] for d in data if d['franja'] != 'Sin Franja')))
    is_saturday_shift = (len(unique_franjas) <= 2) or (_now.weekday() == 5)

    # --- Detectar si ya pasó la primera franja (para apoyo inter-zona) ---
    first_f_end = None
    for fr in unique_franjas:
        fs, fe = parse_franja_hours(fr)
        if fe is not None:
            first_f_end = fe
            break
    if first_f_end is None:
        # fallback: si no se puede parsear, usamos mediodía como proxy
        after_first_franja = current_hour >= 12.0
    else:
        # apoyo proactivo se habilita cuando termina la primera franja (+15 min)
        after_first_franja = current_hour >= (first_f_end + 0.25)

    global MAX_IDEAL_LOAD, MAX_ABSOLUTE_LOAD, ORDER_DURATION_HOURS, MAX_ORDER_DURATION_HOURS
    if is_saturday_shift:
        print(f"MODO SABADO DETECTADO ({len(unique_franjas)} franjas). Ajustando limites: 3-4 ordenes.")
        MAX_IDEAL_LOAD = 3
        MAX_ABSOLUTE_LOAD = 4
        ORDER_DURATION_HOURS = 0.75
        MAX_ORDER_DURATION_HOURS = 1.1
    else:
        MAX_IDEAL_LOAD = 5
        MAX_ABSOLUTE_LOAD = 6   # ✅ LV
        ORDER_DURATION_HOURS = 1.0
        MAX_ORDER_DURATION_HOURS = 1.5

    suggestions = []
    alerts = []
    discarded_suggestions = []

    def add_discard(order, donor, receiver, zone, motivo, extra=""):
        try:
            discarded_suggestions.append({
                "zona": zone,
                "subzona": order.get("subzona", "") if isinstance(order, dict) else "",
                "direccion": order.get("address", "") if isinstance(order, dict) else "",
                "order_id": order.get("order_id", "") if isinstance(order, dict) else "",
                "origen": donor or "",
                "destino": receiver or "",
                "franja": order.get("franja", "") if isinstance(order, dict) else "",
                "motivo": motivo,
                "detalle": extra,
            })
        except Exception:
            pass
    zone_summaries = []
    subzone_summaries = []

    # =========================
    # Cargas y mapas globales
    # =========================
    tech_total = {}
    tech_finalized = {}
    tech_pending = {}
    tech_effective_load = {}
    tech_completion_credit_map = {}
    tech_movable = {}
    tech_all_orders = {}
    tech_locations = {}
    tech_franja_counts = {}
    tech_subzones = {}
    tech_has_near_finish = {}
    tech_main_zone = {}

    # Para optimización por edificio
    active_buildings = {}   # addr_key -> set(techs) con estado >=2 en esa unidad
    building_orders_programmed = {}  # addr_key -> list de programadas

    for d in data:
        t = d['tech']
        status_lower = d['status'].lower()

        # Programadas por unidad
        if d.get('addr_key') and is_status(status_lower, MOVABLE_STATUSES):
            building_orders_programmed.setdefault(d['addr_key'], []).append(d)

        if t == "SIN_ASIGNAR":
            if is_status(status_lower, MOVABLE_STATUSES):
                tech_movable.setdefault(t, []).append(d)
            continue

        tech_all_orders.setdefault(t, []).append(d)
        tech_total[t] = tech_total.get(t, 0) + 1
        tech_effective_load[t] = tech_effective_load.get(t, 0.0) + status_effective_weight(d['status'])
        tech_completion_credit_map[t] = tech_completion_credit_map.get(t, 0.0) + status_completion_credit(d['status'])

        if t not in tech_main_zone and d['zona'] != "SIN_ZONA":
            tech_main_zone[t] = d['zona']

        if is_status(status_lower, FINALIZED_STATUSES):
            tech_finalized[t] = tech_finalized.get(t, 0) + 1
        else:
            tech_pending[t] = tech_pending.get(t, 0) + 1

            tech_franja_counts.setdefault(t, {})
            tech_franja_counts[t][d['franja']] = tech_franja_counts[t].get(d['franja'], 0) + 1

            if d['lat'] != 0 and d['lon'] != 0:
                tech_locations.setdefault(t, []).append((d['lat'], d['lon']))

            tech_subzones.setdefault(t, set()).add(d['subzona'])

            if is_status(status_lower, NEAR_FINISH_STATUSES):
                tech_has_near_finish[t] = True

        if is_status(status_lower, MOVABLE_STATUSES):
            tech_movable.setdefault(t, []).append(d)

        prog = get_status_progress(d['status'])
        if prog >= 2 and prog < 6 and d.get('addr_key'):
            active_buildings.setdefault(d['addr_key'], set()).add(t)

    donors_interzone_count = {}
    receivers_interzone_count = {}
    techs_moved_from_zone = set()

    # Mapa base de técnicos por zona para validaciones inter-zona tempranas
    zone_techs = {}
    for t, zname in tech_main_zone.items():
        if t == "SIN_ASIGNAR" or not zname:
            continue
        zone_techs.setdefault(zname, []).append(t)

    # =========================
    # Analisis por zona (alertas + nivelación principal)
    # =========================
    for z in zones:
        zone_data = [d for d in data if d['zona'] == z]

        all_techs_in_zone = sorted(set(d['tech'] for d in zone_data if d['tech'] != "SIN_ASIGNAR"))
        active_techs_in_zone = [t for t in all_techs_in_zone if tech_pending.get(t, 0) > 0 or tech_finalized.get(t, 0) > 0]

        # ===== Alertas =====
        for t in active_techs_in_zone:
            total = tech_total.get(t, 0)

            if total > MAX_ABSOLUTE_LOAD:
                alerts.append({
                    'tipo': 'SOBRECARGA',
                    'zona': z, 'tecnico': t,
                    'detalle': f"Tiene {total} ordenes totales (max: {MAX_ABSOLUTE_LOAD})"
                })

            all_tech_orders = tech_all_orders.get(t, [])
            active_conflicts = [o for o in all_tech_orders if 2 <= get_status_progress(o['status']) < 6]
            if len(active_conflicts) > 1:
                ids = [str(o['order_id']) for o in active_conflicts]
                alerts.append({
                    'tipo': 'MULTI-ESTADO ACTIVO',
                    'zona': z, 'tecnico': t,
                    'detalle': f"El técnico tiene {len(active_conflicts)} órdenes activas simultáneas: {', '.join(ids)}."
                })

            # Alertas riesgo franja
            active_list = [o for o in all_tech_orders if 1 <= get_status_progress(o['status']) < 6]
            movable_list = sorted(
                [o for o in all_tech_orders if is_status(o['status'].lower(), MOVABLE_STATUSES)],
                key=lambda x: parse_franja_hours(x['franja'])[0] if parse_franja_hours(x['franja'])[0] is not None else 0
            )

            proj_normal = current_hour
            proj_max = current_hour

            if active_list:
                o_act = active_list[0]
                rem_n = estimate_remaining_hours(o_act['status'], o_act.get('onsite_hour'), current_hour)
                rem_m = rem_n * (MAX_ORDER_DURATION_HOURS / ORDER_DURATION_HOURS) if ORDER_DURATION_HOURS > 0 else rem_n
                proj_normal += rem_n
                proj_max += rem_m

            for dmv in movable_list:
                f_start, f_end = parse_franja_hours(dmv['franja'])
                if f_start is None:
                    continue

                arrival_normal = max(proj_normal, f_start)
                arrival_max = max(proj_max, f_start)

                proj_normal = arrival_normal + ORDER_DURATION_HOURS
                proj_max = arrival_max + MAX_ORDER_DURATION_HOURS

                if f_end is not None and arrival_max > f_end:
                    alerts.append({
                        'tipo': 'FRANJA EN RIESGO',
                        'zona': z, 'tecnico': t,
                        'detalle': f"Orden {dmv['order_id']} (Franja {dmv['franja']}) - Arribo ~{format_hour(arrival_max)} (Límite {format_hour(f_end)})."
                    })

            # Duplicadas existentes
            if t in tech_franja_counts:
                duplicated = count_duplicated_slots(tech_franja_counts[t])
                if duplicated > MAX_DUPLICATED_SLOTS:
                    franjas_dup = [f"{f}: {c}" for f, c in tech_franja_counts[t].items() if c >= 2]
                    alerts.append({
                        'tipo': 'FRANJAS DUPLICADAS',
                        'zona': z, 'tecnico': t,
                        'detalle': f"{duplicated} franjas duplicadas: {', '.join(franjas_dup)}"
                    })

        # ===== Resumen zona =====
        active_techs_in_zone = [t for t, m_zone in tech_main_zone.items() if m_zone == z]
        all_techs_in_zone = active_techs_in_zone

        unassigned_orders = [
            d for d in zone_data if d['tech'] == "SIN_ASIGNAR"
            and is_status(d['status'].lower(), MOVABLE_STATUSES)
        ]

        initial_pending_total = sum(tech_pending.get(t, 0) for t in active_techs_in_zone)
        initial_finalized_total = sum(tech_finalized.get(t, 0) for t in active_techs_in_zone)
        zone_pends = [tech_pending.get(t, 0) for t in active_techs_in_zone]

        # Alerta de bajo avance relativo: técnico con solo 0-1 orden finalizada/por auditar
        # mientras varios compañeros de la misma zona ya van en 2-3 o más.
        zone_completed_counts = {t: tech_finalized.get(t, 0) for t in active_techs_in_zone if t != "SIN_ASIGNAR"}
        for t, completed_count in zone_completed_counts.items():
            peer_counts = [c for ot, c in zone_completed_counts.items() if ot != t]
            if should_alert_low_progress(t, completed_count, peer_counts):
                peer_avg = round(sum(peer_counts) / len(peer_counts), 1) if peer_counts else 0
                top_peers = sorted(peer_counts, reverse=True)[:3]
                alerts.append({
                    'tipo': 'BAJO AVANCE RELATIVO',
                    'zona': z,
                    'tecnico': t,
                    'detalle': (
                        f"Lleva {completed_count} orden(es) finalizada(s)/por auditar, mientras otros técnicos de la zona ya van en "
                        f"{', '.join(str(x) for x in top_peers)} (promedio pares: {peer_avg}). "
                        f"Validar con supervisor posible baja productividad o sobrecarga operativa al equipo."
                    )
                })

        current_zone_summary = {
            'zona': z,
            'techs': len(active_techs_in_zone),
            'pendientes_inicial': initial_pending_total,
            'pendientes_final': initial_pending_total,
            'sin_asignar_inicial': len(unassigned_orders),
            'sin_asignar_final': len(unassigned_orders),
            'total_finalizadas': initial_finalized_total,
            'avg_inicial': round(float(initial_pending_total) / len(active_techs_in_zone), 1) if active_techs_in_zone else 0.0,
            'min_inicial': min(zone_pends) if zone_pends else 0,
            'max_inicial': max(zone_pends) if zone_pends else 0,
        }
        zone_summaries.append(current_zone_summary)

        # ===== Resumen por técnico (subzonas/estados) =====
        active_zone_pending_data = [d for d in zone_data if not is_status(d['status'].lower(), FINALIZED_STATUSES)]

        tech_subzone_map = {}
        for dd in active_zone_pending_data:
            t = dd['tech']
            sz = dd['subzona']
            st = dd['status']
            tech_subzone_map.setdefault(t, {}).setdefault(sz, {})
            tech_subzone_map[t][sz][st] = tech_subzone_map[t][sz].get(st, 0) + 1

        # ✅ Incluir técnicos sin pendientes (pendientes=0) pero con finalizadas en esta zona
        # Para que la hoja "Distribucion por Tecnico" muestre también quién ya terminó todo.
        techs_in_zone_all = sorted(set(d['tech'] for d in zone_data if d['tech'] != "SIN_ASIGNAR"))
        for t0 in techs_in_zone_all:
            if tech_pending.get(t0, 0) == 0 and tech_finalized.get(t0, 0) > 0:
                tech_subzone_map.setdefault(t0, {})

        for t in sorted(tech_subzone_map.keys()):

            sz_map = tech_subzone_map[t]
            total_pending = sum(sum(v for v in sts.values()) for sts in sz_map.values())
            subzona_lines = []
            if not sz_map:
                subzona_lines.append("SIN PENDIENTES")
            for sz in sorted(sz_map.keys()):

                sts = sz_map[sz]
                estados_str = "  /  ".join(f"{k}({v})" for k, v in sts.items())
                subzona_lines.append(f"{sz}: {estados_str}")
            subzone_summaries.append({
                'zona': z,
                'tecnico': t,
                'subzonas_detalle': "\n".join(subzona_lines),
                'num_subzonas': len(sz_map),
                'pendientes': total_pending,
                'finalizadas': tech_finalized.get(t, 0) if t != "SIN_ASIGNAR" else 0,
                'carga_total': round(tech_effective_load.get(t, tech_total.get(t, 0) if t != "SIN_ASIGNAR" else total_pending), 2),
            })

        # =========================
        # PASO 5: NIVELACION por carga y riesgo (tu lógica, casi igual)
        # =========================
        donors = []

        unassigned_in_zone = [d for d in tech_movable.get("SIN_ASIGNAR", []) if d['zona'] == z]
        if unassigned_in_zone:
            donors.append("SIN_ASIGNAR")

        avg_zone_pending = (sum(tech_effective_load.get(t, tech_pending.get(t, 0)) for t in active_techs_in_zone) / len(active_techs_in_zone)) if active_techs_in_zone else 0

        for t in sorted(active_techs_in_zone):
            if t == "SIN_ASIGNAR":
                continue

            if tech_total.get(t, 0) > MAX_ABSOLUTE_LOAD and t in tech_movable and tech_movable[t]:
                if t not in donors: donors.append(t)
                continue

            if (tech_effective_load.get(t, tech_pending.get(t, 0)) > MAX_IDEAL_LOAD or tech_effective_load.get(t, tech_pending.get(t, 0)) > (avg_zone_pending + 0.8)) \
               and t in tech_movable and tech_movable[t]:
                if t not in donors: donors.append(t)
                continue

        for t in active_techs_in_zone:
            if t == "SIN_ASIGNAR":
                continue
            for dmv in tech_movable.get(t, []):
                f_start, f_end = parse_franja_hours(dmv['franja'])
                if f_start is None:
                    continue
                active_order = next((o for o in tech_all_orders.get(t, []) if 1 <= get_status_progress(o['status']) < 6), None)
                rem_h = estimate_remaining_hours(active_order['status'], active_order.get('onsite_hour'), current_hour) if active_order else 0
                prog_before = sum(
                    1 for o in tech_all_orders.get(t, [])
                    if is_status(o['status'].lower(), MOVABLE_STATUSES)
                    and parse_franja_hours(o['franja'])[0] is not None
                    and parse_franja_hours(o['franja'])[0] < f_start
                )
                est_ready = current_hour + rem_h + (prog_before * ORDER_DURATION_HOURS)

                if f_end is not None and est_ready > (f_end + 0.1):
                    if t not in donors: donors.append(t)
                    break

        def has_zone_capacity(zone_name):
            techs_in_z = [t for t, mz in tech_main_zone.items() if mz == zone_name]
            active_now = sum(1 for t in techs_in_z if t not in techs_moved_from_zone)
            return active_now > 1

        for donor in donors:
            if donor == "SIN_ASIGNAR":
                donor_orders = [d for d in tech_movable.get("SIN_ASIGNAR", []) if d['zona'] == z]
            else:
                donor_orders = list(tech_movable.get(donor, []))

            if not donor_orders:
                continue

            donor_orders.sort(key=lambda x: parse_franja_hours(x['franja'])[0] or 99)

            if donor == "SIN_ASIGNAR":
                moves_limit = len(donor_orders)
            else:
                donor_eff = tech_effective_load.get(donor, tech_pending.get(donor, 0))
                moves_limit = max(1, math.ceil(max(0.0, donor_eff - MAX_IDEAL_LOAD)))

            moved_count = 0
            for order in list(donor_orders):
                if moved_count >= moves_limit:
                    break

                best_receiver = None
                best_score = float('inf')
                best_detail = ""

                for pass_num in [1, 2]:
                    if best_receiver:
                        break

                    if pass_num == 1:
                        recipients = [r for r in active_techs_in_zone if r != donor]
                    else:
                        if donors_interzone_count.get(donor, 0) >= 1:
                            break
                        allowed_neighbor_zones = ZONE_ADJACENCY.get(z.upper(), [])
                        recipients = [
                            t for t in tech_total
                            if t not in all_techs_in_zone
                            and t != donor and t != "SIN_ASIGNAR"
                            and tech_main_zone.get(t, "").upper() in allowed_neighbor_zones
                        ]

                    for r in recipients:
                        if tech_total.get(r, 0) >= MAX_ABSOLUTE_LOAD:
                            add_discard(order, donor, r, z, "destino_excede_maximo", f"total={tech_total.get(r, 0)}")
                            continue

                        donor_eff = tech_effective_load.get(donor, tech_pending.get(donor, 0)) if donor != "SIN_ASIGNAR" else None
                        recv_eff = tech_effective_load.get(r, tech_pending.get(r, 0))
                        move_ok, move_reason, savings_km, donor_dist, receiver_dist, load_gap = should_allow_move_by_balance_or_route(
                            order, donor, r, donor_eff, recv_eff, tech_pending, tech_total, tech_all_orders, tech_locations, current_hour,
                            tech_franja_counts, tech_subzones, tech_completion_credit_map, tech_effective_load
                        ) if donor != "SIN_ASIGNAR" else (True, "sin_asignar", None, None, None, None)
                        if donor != "SIN_ASIGNAR" and not move_ok:
                            add_discard(order, donor, r, z, "sin_beneficio_real", move_reason)
                            continue

                        can_handle, handle_reason = can_tech_handle_franja(
                            tech_franja_counts.get(r, {}),
                            tech_all_orders.get(r, []),
                            order['franja'],
                            current_hour
                        )
                        if not can_handle:
                            add_discard(order, donor, r, z, "franja_no_viable", handle_reason)
                            continue

                        zone_only_no_coords = False
                        if allow_zone_only_assignment_when_no_coords(order):
                            if not allow_subzone_move_when_no_coords(order, r, tech_subzones, tech_all_orders):
                                zone_only_no_coords = True
                        elif not allow_subzone_move_when_no_coords(order, r, tech_subzones, tech_all_orders):
                            add_discard(order, donor, r, z, "sin_coords_subzona_no_valida", "solo misma unidad o misma subzona")
                            continue

                        is_it_interzone = (tech_main_zone.get(r) != z)
                        zone_rank = zone_rank_from(z, tech_main_zone.get(r)) if is_it_interzone else 0
                        if is_it_interzone:
                            if donor in techs_moved_from_zone:
                                add_discard(order, donor, r, z, "donante_ya_salio_de_zona", "max 1 cruce")
                                continue
                            if receivers_interzone_count.get(r, 0) >= MAX_INTERZONE_ASSIGNMENTS_PER_TECH:
                                add_discard(order, donor, r, z, "destino_interzona_al_limite", "max 1 apoyo")
                                continue
                            if zone_rank != 1:
                                add_discard(order, donor, r, z, "zona_no_vecina", f"rank={zone_rank}")
                                continue
                            if not has_zone_capacity(z):
                                add_discard(order, donor, r, z, "origen_sin_capacidad", "debe quedar capacidad minima")
                                continue
                            inter_ok, inter_reason = should_offer_interzone_support(z, tech_main_zone.get(r), donor, zone_techs, tech_pending, tech_all_orders, tech_effective_load, receivers_interzone_count)
                            if not inter_ok:
                                add_discard(order, donor, r, z, "interzona_no_conviene", inter_reason)
                                continue

                        score = tech_load_score(r, tech_pending, tech_total, tech_effective_load, tech_completion_credit_map, tech_franja_counts, tech_subzones, tech_all_orders)

                        dist = estimate_distance_to_receiver(order, r, tech_all_orders, tech_locations)
                        if dist is not None:
                            if dist > MAX_ALLOWED_DISTANCE_KM:
                                add_discard(order, donor, r, z, "distancia_excesiva", f"dist={dist:.2f}km")
                                continue
                            score += dist * 300
                        else:
                            dist = 0.0

                        route_bonus = 0
                        if donor != "SIN_ASIGNAR":
                            if savings_km is not None:
                                route_bonus = min(2600, savings_km * 900)
                                score -= route_bonus
                            elif donor_dist is not None and receiver_dist is not None and receiver_dist < donor_dist:
                                route_bonus = min(1800, (donor_dist - receiver_dist) * 700)
                                score -= route_bonus

                        bonus_sub = 0
                        if order['subzona'] in tech_subzones.get(r, set()):
                            bonus_sub = -2400
                            score += bonus_sub

                        if order.get('addr_key') and order['addr_key'] in set(o.get('addr_key') for o in tech_all_orders.get(r, [])):
                            score -= 2800

                        if order['subzona'] not in tech_subzones.get(r, set()) and len(tech_subzones.get(r, set())) >= MAX_SUBZONES_SOFT:
                            score += FRAGMENTATION_PENALTY * 1.2

                        arr_n, arr_m, f_start, f_end = estimate_arrival_for_franja(
                            tech_all_orders.get(r, []), order['franja'], current_hour
                        )
                        late_pen = 0
                        if f_end is not None and arr_m is not None and arr_m > f_end:
                            late_hours = (arr_m - f_end)
                            late_pen = late_hours * 12000
                            score += late_pen
                        elif f_start is not None and arr_n is not None and arr_n > f_start:
                            score += (arr_n - f_start) * 1200

                        if zone_only_no_coords:
                            score += ZONE_ONLY_NO_COORDS_PENALTY
                        if is_it_interzone:
                            score += INTERZONE_DISTANCE_PENALTY * zone_rank

                        if score < best_score:
                            best_score = score
                            best_receiver = r
                            best_detail = f"score={score:.0f} | motivo={move_reason} | gap={load_gap if load_gap is not None else 'NA'} | ahorro_km={(savings_km if savings_km is not None else 0):.2f} | eff_load={recv_eff:.2f} | total={tech_total.get(r,0)} | dist={dist:.2f} | route_bonus={route_bonus:.0f} | bonus_sub={bonus_sub} | late_pen={late_pen:.0f} | {handle_reason}"

                    if pass_num == 2 and best_receiver:
                        donors_interzone_count[donor] = donors_interzone_count.get(donor, 0) + 1
                        receivers_interzone_count[best_receiver] = receivers_interzone_count.get(best_receiver, 0) + 1
                        techs_moved_from_zone.add(donor)

                if best_receiver:
                    dist_km = 0.0
                    est_dist = estimate_distance_to_receiver(order, best_receiver, tech_all_orders, tech_locations)
                    if est_dist is not None:
                        dist_km = est_dist

                    is_interzone = tech_main_zone.get(best_receiver) != z
                    suggestions.append({
                        'zona': f"{z} (AYUDA EXTERNA)" if is_interzone else z,
                        'subzona': order['subzona'],
                        'origen': donor,
                        'destino': best_receiver,
                        'order_id': order['order_id'],
                        'franja': order['franja'],
                        'address': order.get('address', ''),
                        'distancia_estimada': f"{dist_km:.2f} km",
                        'alerta': f"Inter-Zona ({tech_main_zone.get(best_receiver)})" if is_interzone else ("Optimización Ruta / Carga" if donor != "SIN_ASIGNAR" else "Sin Asignar"),
                        'pendientes_origen': tech_pending.get(donor, 0),
                        'pendientes_destino': tech_pending.get(best_receiver, 0),
                        'justificacion': best_detail
                    })

                    if donor != "SIN_ASIGNAR":
                        tech_total[donor] -= 1
                        tech_pending[donor] = max(0, tech_pending.get(donor, 0) - 1)
                        tech_effective_load[donor] = max(0.0, tech_effective_load.get(donor, 0.0) - status_effective_weight(order['status']))
                        current_zone_summary['pendientes_final'] -= 1
                    else:
                        current_zone_summary['sin_asignar_final'] -= 1
                        current_zone_summary['pendientes_final'] += 1

                    tech_total[best_receiver] = tech_total.get(best_receiver, 0) + 1
                    tech_pending[best_receiver] = tech_pending.get(best_receiver, 0) + 1
                    tech_effective_load[best_receiver] = tech_effective_load.get(best_receiver, 0.0) + status_effective_weight(order['status'])

                    tech_franja_counts.setdefault(best_receiver, {})
                    tech_franja_counts[best_receiver][order['franja']] = tech_franja_counts[best_receiver].get(order['franja'], 0) + 1

                    tech_all_orders.setdefault(best_receiver, []).append(order)

                    donor_orders.remove(order)
                    tech_movable[donor].remove(order)
                    moved_count += 1
                else:
                    add_discard(order, donor, '', z, 'sin_destino_viable', 'ningun receptor cumplio reglas duras')

        # desbalance final
        if active_techs_in_zone:
            final_pends = [tech_pending.get(t, 0) for t in active_techs_in_zone]
            current_zone_summary['desbalance_final'] = max(final_pends) - min(final_pends)
        else:
            current_zone_summary['desbalance_final'] = 0

    # ==========================================================
    # PASO 5.6: ASIGNAR "POR PROGRAMAR" (SIN_ASIGNAR) dentro de la misma zona
    # Objetivo: garantizar que TODA orden sin técnico (ej. status "Por programar")
    # quede sugerida a un técnico de la zona/subzona, si existe capacidad.
    # - Prioriza misma subzona / misma unidad (addr_key)
    # - Respeta franja: 2 por franja y 1 franja duplicada (con excepción condicional por misma unidad)
    # ==========================================================
    remaining_unassigned = list(tech_movable.get("SIN_ASIGNAR", []))
    if remaining_unassigned:
        # construir mapa técnicos por zona desde tech_main_zone (más confiable que zone loop anterior)
        zone_techs_all = {}
        for t, mz in tech_main_zone.items():
            if t == "SIN_ASIGNAR" or not mz:
                continue
            zone_techs_all.setdefault(mz, []).append(t)

        def receiver_score(order, receiver):
            score = tech_load_score(receiver, tech_pending, tech_total, tech_effective_load, tech_completion_credit_map, tech_franja_counts, tech_subzones, tech_all_orders)

            if order.get('subzona') and order['subzona'] in tech_subzones.get(receiver, set()):
                score -= 2000

            if order.get('addr_key') and order['addr_key'] in set(o.get('addr_key') for o in tech_all_orders.get(receiver, [])):
                score -= 2500

            dist = estimate_distance_to_receiver(order, receiver, tech_all_orders, tech_locations)
            if dist is not None:
                if dist > MAX_ALLOWED_DISTANCE_KM:
                    return float('inf')
                score += dist * 300
                score -= min(2200, max(0, (MAX_ALLOWED_DISTANCE_KM - dist)) * 180)
            return score

        for order in list(remaining_unassigned):
            z = order.get('zona')
            if not z or z == "SIN_ZONA":
                continue

            # receptores dentro de la misma zona
            recipients = zone_techs_all.get(z, [])
            if not recipients:
                continue

            best_r = None
            best_s = float('inf')
            best_reason = ""

            for r in recipients:
                if tech_total.get(r, 0) >= MAX_ABSOLUTE_LOAD:
                    continue

                same_unit_flag = False
                # misma unidad por addr_key (si hay)
                if order.get('addr_key'):
                    same_unit_flag = order['addr_key'] in set(o.get('addr_key') for o in tech_all_orders.get(r, []) if o.get('addr_key'))

                can_ok, reason = can_tech_handle_franja(
                    tech_franja_counts.get(r, {}),
                    tech_all_orders.get(r, []),
                    order.get('franja', ''),
                    current_hour,
                    allow_same_unit_override=True,
                    same_unit=same_unit_flag
                )
                if not can_ok:
                    continue

                s = receiver_score(order, r)
                zone_only_no_coords = False
                if allow_zone_only_assignment_when_no_coords(order):
                    # si no cumple misma unidad/subzona, lo dejamos como candidato pero penalizado y marcado
                    if not allow_subzone_move_when_no_coords(order, r, tech_subzones, tech_all_orders):
                        zone_only_no_coords = True
                s = receiver_score(order, r)
                if zone_only_no_coords:
                    s += ZONE_ONLY_NO_COORDS_PENALTY  # penalización fuerte: solo se elige si no hay mejores
                    reason = (reason + ' | SIN COORDS: sugerencia por zona (no por subzona)')
                if s < best_s:
                    best_s = s
                    best_r = r
                    best_reason = reason

            if not best_r:
                continue

            # Registrar sugerencia
            dist_msg = "N/A"
            est_dist = estimate_distance_to_receiver(order, best_r, tech_all_orders, tech_locations)
            if est_dist is not None:
                dist_msg = f"{est_dist:.2f} km"

            alert = "ASIGNAR POR PROGRAMAR"
            if allow_zone_only_assignment_when_no_coords(order) and ("SIN COORDS" in best_reason):
                alert = "ASIGNAR POR ZONA (SIN COORDS)"
            if "EXCEPCION" in best_reason:
                alert = "ASIGNAR POR PROGRAMAR (CONDICIONAL: MISMA UNIDAD)"

            suggestions.append({
                'zona': z,
                'subzona': order.get('subzona', ''),
                'origen': "SIN_ASIGNAR",
                'destino': best_r,
                'order_id': order.get('order_id', ''),
                'franja': order.get('franja', ''),
                'address': order.get('address', ''),
                'distancia_estimada': dist_msg if dist_msg else "N/A",
                'alerta': alert,
                'pendientes_origen': 0,
                'pendientes_destino': tech_pending.get(best_r, 0),
                'justificacion': f"Orden sin técnico (Por programar). Asignación sugerida en zona {z}. {best_reason}."
            })

            # Actualizar contadores internos (simulación)
            tech_total[best_r] = tech_total.get(best_r, 0) + 1
            tech_pending[best_r] = tech_pending.get(best_r, 0) + 1
            tech_effective_load[best_r] = tech_effective_load.get(best_r, 0.0) + status_effective_weight(order['status'])
            tech_franja_counts.setdefault(best_r, {})
            fr = order.get('franja', '')
            tech_franja_counts[best_r][fr] = tech_franja_counts[best_r].get(fr, 0) + 1
            tech_all_orders.setdefault(best_r, []).append(order)

            # remover del pool SIN_ASIGNAR para evitar duplicados
            try:
                tech_movable.get("SIN_ASIGNAR", []).remove(order)
            except:
                pass



    # ==========================================================
    # PASO 5.9: APOYO PROACTIVO INTER-ZONA (post 1ra franja)
    # - Si una zona vecina está saturada y esta zona está liviana,
    #   sugerimos mover capacidad (técnico libre) a la zona vecina.
    # - Regla: el técnico a mover debe quedar SIN pendientes, o
    #   redistribuimos sus programadas a otros técnicos de su zona.
    # ==========================================================
    if after_first_franja:
        # --- construir mapa de técnicos por zona y avg de pendientes ---
        zone_techs = {}
        for t, zname in tech_main_zone.items():
            if not zname or t == "SIN_ASIGNAR":
                continue
            zone_techs.setdefault(zname, []).append(t)

        zone_avg_pending = {}
        for zname, techs in zone_techs.items():
            if not techs:
                zone_avg_pending[zname] = 0.0
            else:
                zone_avg_pending[zname] = sum(tech_pending.get(t, 0) for t in techs) / len(techs)

        # --- addr_key por técnico (para reasignar sin coords por misma unidad) ---
        tech_addr_keys = {}
        for tt, orders in tech_all_orders.items():
            keys = set()
            for o in orders:
                ak = o.get('addr_key')
                if ak:
                    keys.add(ak)
            tech_addr_keys[tt] = keys

        def tech_has_active_order(tname):
            # cualquier estado activo (inbound/en sitio/iniciado...) cuenta como NO movible de zona
            return any(1 <= get_status_progress(o['status']) < 6 for o in tech_all_orders.get(tname, []))

        def redistribute_to_free_tech(tname, zname):
            # Redistribuye órdenes programadas del técnico dentro de su zona para dejarlo sin pendientes.
            movable_list = [o for o in list(tech_movable.get(tname, [])) if o.get('zona') == zname]
            if not movable_list:
                return True

            recipients = [r for r in zone_techs.get(zname, []) if r != tname]
            if not recipients:
                return False

            def fr_start(od):
                fs, _ = parse_franja_hours(od.get('franja'))
                return fs if fs is not None else 99
            movable_list.sort(key=fr_start)

            for order in list(movable_list):
                best_r = None
                best_score = float('inf')
                best_reason = ""

                for r in recipients:
                    if tech_total.get(r, 0) >= MAX_ABSOLUTE_LOAD:
                        continue

                    can_ok, reason = can_tech_handle_franja(
                        tech_franja_counts.get(r, {}),
                        tech_all_orders.get(r, []),
                        order.get('franja', ''),
                        current_hour,
                        allow_same_unit_override=False,
                        same_unit=False
                    )
                    if not can_ok:
                        continue

                    dist = 0.0
                    if order.get('lat', 0) and order.get('lon', 0) and r in tech_locations:
                        dist = estimate_distance_to_receiver(order, r, tech_all_orders, tech_locations)
                        if dist is None or dist > MAX_ALLOWED_DISTANCE_KM:
                            continue
                    else:
                        ak = order.get('addr_key') or ""
                        if not ak or ak not in tech_addr_keys.get(r, set()):
                            continue

                    score = tech_load_score(r, tech_pending, tech_total, tech_effective_load, tech_completion_credit_map, tech_franja_counts, tech_subzones, tech_all_orders) + dist * 300
                    if order.get('subzona') in tech_subzones.get(r, set()):
                        score -= 1500

                    if score < best_score:
                        best_score = score
                        best_r = r
                        best_reason = reason

                if not best_r:
                    return False

                suggestions.append({
                    'zona': zname,
                    'subzona': order.get('subzona', ''),
                    'origen': tname,
                    'destino': best_r,
                    'order_id': order.get('order_id', ''),
                    'franja': order.get('franja', ''),
                    'address': order.get('address', ''),
                    'distancia_estimada': "N/A" if not (order.get('lat',0) and order.get('lon',0)) else "Ver distancia",
                    'alerta': "LIBERAR TECNICO (PRE-APOYO)",
                    'pendientes_origen': tech_pending.get(tname, 0),
                    'pendientes_destino': tech_pending.get(best_r, 0),
                    'justificacion': f"Redistribución para liberar técnico y apoyar otra zona. {best_reason}."
                })

                tech_total[tname] = max(0, tech_total.get(tname, 0) - 1)
                tech_pending[tname] = max(0, tech_pending.get(tname, 0) - 1)
                tech_effective_load[tname] = max(0.0, tech_effective_load.get(tname, 0.0) - status_effective_weight(order['status']))
                if tname in tech_franja_counts:
                    cnt = tech_franja_counts[tname].get(order.get('franja',''), 0)
                    if cnt <= 1:
                        tech_franja_counts[tname].pop(order.get('franja',''), None)
                    else:
                        tech_franja_counts[tname][order.get('franja','')] = cnt - 1

                tech_total[best_r] = tech_total.get(best_r, 0) + 1
                tech_pending[best_r] = tech_pending.get(best_r, 0) + 1
                tech_effective_load[best_r] = tech_effective_load.get(best_r, 0.0) + status_effective_weight(order['status'])
                tech_franja_counts.setdefault(best_r, {})
                tech_franja_counts[best_r][order.get('franja','')] = tech_franja_counts[best_r].get(order.get('franja',''), 0) + 1

                try:
                    tech_all_orders.get(tname, []).remove(order)
                except:
                    pass
                tech_all_orders.setdefault(best_r, []).append(order)

                if tname in tech_movable and order in tech_movable[tname]:
                    tech_movable[tname].remove(order)
                tech_movable.setdefault(best_r, []).append(order)

                order['tech'] = best_r

            return tech_pending.get(tname, 0) == 0

        # --- ejecutar apoyo proactivo ---
        for z_low in zones:
            low_avg = zone_avg_pending.get(z_low, 0.0)
            if low_avg > (MAX_IDEAL_LOAD - 2):
                continue

            neighbors = ZONE_ADJACENCY.get(z_low.upper(), [])
            if not neighbors:
                continue

            z_high = None
            best_diff = 0.0
            for nz in neighbors:
                high_avg = zone_avg_pending.get(nz, None)
                if high_avg is None:
                    continue
                diff = high_avg - low_avg
                if high_avg >= MAX_IDEAL_LOAD and diff >= 2.0 and diff > best_diff:
                    best_diff = diff
                    z_high = nz

            if not z_high:
                continue

            cand = [t for t in zone_techs.get(z_low, []) if t not in techs_moved_from_zone]
            if len(cand) <= 1:
                continue

            free_cand = [t for t in cand if tech_pending.get(t,0) == 0 and not tech_has_active_order(t)]
            donor = None
            donor_freed_by = "YA LIBRE"
            if free_cand:
                donor = free_cand[0]
            else:
                cand2 = [t for t in cand if not tech_has_active_order(t)]
                if cand2:
                    cand2.sort(key=lambda t: tech_pending.get(t,0))
                    for t_try in cand2[:3]:
                        if redistribute_to_free_tech(t_try, z_low):
                            donor = t_try
                            donor_freed_by = "REDISTRIBUCION"
                            break

            if not donor:
                continue

            target = None
            source_label = None
            pool_un = [d for d in tech_movable.get("SIN_ASIGNAR", []) if d.get('zona') == z_high]
            if pool_un:
                pool_un.sort(key=lambda x: (parse_franja_hours(x.get('franja'))[0] or 99))
                target = pool_un[0]
                source_label = "SIN_ASIGNAR"
            else:
                high_techs = zone_techs.get(z_high, [])
                high_techs.sort(key=lambda t: tech_pending.get(t,0), reverse=True)
                for ht in high_techs:
                    if ht == "SIN_ASIGNAR":
                        continue
                    if tech_pending.get(ht,0) < MAX_IDEAL_LOAD:
                        continue
                    cand_orders = [o for o in tech_movable.get(ht, []) if o.get('zona') == z_high]
                    if cand_orders:
                        cand_orders.sort(key=lambda x: (parse_franja_hours(x.get('franja'))[0] or 99))
                        target = cand_orders[0]
                        source_label = ht
                        break

            if target:
                can_ok, reason = can_tech_handle_franja(
                    tech_franja_counts.get(donor, {}),
                    tech_all_orders.get(donor, []),
                    target.get('franja',''),
                    current_hour,
                    allow_same_unit_override=False,
                    same_unit=False
                )
                inter_ok, inter_reason = should_offer_interzone_support(z_low, z_high, donor, zone_techs, tech_pending, tech_all_orders, tech_effective_load, receivers_interzone_count)
                if can_ok and inter_ok:
                    suggestions.append({
                        'zona': f"{z_high} (APOYO PROACTIVO)",
                        'subzona': target.get('subzona',''),
                        'origen': donor,
                        'destino': f"{source_label} (Zona {z_high})",
                        'order_id': target.get('order_id',''),
                        'franja': target.get('franja',''),
                        'address': target.get('address',''),
                        'distancia_estimada': "N/A" if not (target.get('lat',0) and target.get('lon',0)) else "Ver distancia",
                        'alerta': "APOYO PROACTIVO INTERZONA",
                        'pendientes_origen': tech_pending.get(donor, 0),
                        'pendientes_destino': round(zone_avg_pending.get(z_high,0.0), 1),
                        'justificacion': (
                            f"Post 1ra franja: {z_low} liviana (avg {low_avg:.1f}) y {z_high} saturada (avg {zone_avg_pending.get(z_high,0):.1f}). "
                            f"Técnico {donor} {donor_freed_by}. {reason}."
                        )
                    })
                    techs_moved_from_zone.add(donor)
                    receivers_interzone_count[donor] = receivers_interzone_count.get(donor, 0) + 1

                    if source_label == "SIN_ASIGNAR":
                        try:
                            tech_movable.get("SIN_ASIGNAR", []).remove(target)
                        except:
                            pass
                    else:
                        try:
                            tech_movable.get(source_label, []).remove(target)
                        except:
                            pass
                else:
                    add_discard(target, donor, f"CAPACIDAD STANDBY ({z_high})", z_high, "apoyo_interzona_descartado", (reason + '; ' + inter_reason) if 'inter_reason' in locals() else reason)
                    suggestions.append({
                        'zona': f"{z_high} (APOYO PROACTIVO)",
                        'subzona': "",
                        'origen': donor,
                        'destino': f"CAPACIDAD STANDBY (Zona {z_high})",
                        'order_id': "N/A",
                        'franja': "Post 1ra franja",
                        'address': "",
                        'distancia_estimada': "",
                        'alerta': "APOYO PROACTIVO INTERZONA",
                        'pendientes_origen': tech_pending.get(donor, 0),
                        'pendientes_destino': round(zone_avg_pending.get(z_high,0.0), 1),
                        'justificacion': (
                            f"Post 1ra franja: {z_low} liviana (avg {low_avg:.1f}) y {z_high} saturada (avg {zone_avg_pending.get(z_high,0):.1f}). "
                            f"Técnico {donor} {donor_freed_by}. No se asigna orden específica: {reason if not can_ok else inter_reason}."
                        )
                    })
                    techs_moved_from_zone.add(donor)
                    receivers_interzone_count[donor] = receivers_interzone_count.get(donor, 0) + 1
            else:
                suggestions.append({
                    'zona': f"{z_high} (APOYO PROACTIVO)",
                    'subzona': "",
                    'origen': donor,
                    'destino': f"CAPACIDAD STANDBY (Zona {z_high})",
                    'order_id': "N/A",
                    'franja': "Post 1ra franja",
                    'address': "",
                    'distancia_estimada': "",
                    'alerta': "APOYO PROACTIVO INTERZONA",
                    'pendientes_origen': tech_pending.get(donor, 0),
                    'pendientes_destino': round(zone_avg_pending.get(z_high,0.0), 1),
                    'justificacion': (
                        f"Post 1ra franja: {z_low} liviana (avg {low_avg:.1f}) y {z_high} saturada (avg {zone_avg_pending.get(z_high,0):.1f}). "
                        f"Técnico {donor} {donor_freed_by}. Sin orden concreta; se propone apoyo si llega nueva."
                    )
                })
                techs_moved_from_zone.add(donor)
                receivers_interzone_count[donor] = receivers_interzone_count.get(donor, 0) + 1
    # ==========================================================
    # PASO 6: OPTIMIZACIÓN DE RUTA / UNIDAD + PROXIMIDAD + SWAPS
    # ==========================================================
    all_techs_global = sorted([t for t in tech_total.keys() if t != "SIN_ASIGNAR"])
    PROXIMITY_GAIN_MIN_KM = 1.5

    # 6A: Reasignación por proximidad (intra-zona) - igual que antes
    for t1 in all_techs_global:
        if not tech_movable.get(t1):
            continue
        for t2 in all_techs_global:
            if t2 <= t1:
                continue
            if not tech_movable.get(t2):
                continue

            if tech_main_zone.get(t1) != tech_main_zone.get(t2):
                continue

            if abs(tech_pending.get(t1, 0) - tech_pending.get(t2, 0)) > 1:
                continue

            c1 = get_tech_reference_point(t1, tech_all_orders, tech_locations)
            c2 = get_tech_reference_point(t2, tech_all_orders, tech_locations)

            for o1 in list(tech_movable[t1]):
                # si no hay coords, solo sugerimos si es misma unidad (addr_key) con un activo (lo cubre 6.0 fuera)
                if o1['lat'] == 0:
                    continue

                d_o1_t1 = haversine(o1['lat'], o1['lon'], c1[0], c1[1])
                d_o1_t2 = haversine(o1['lat'], o1['lon'], c2[0], c2[1])
                gain = d_o1_t1 - d_o1_t2

                if gain >= PROXIMITY_GAIN_MIN_KM:
                    can_handle, _ = can_tech_handle_franja(
                        tech_franja_counts.get(t2, {}),
                        tech_all_orders.get(t2, []),
                        o1['franja'], current_hour
                    )
                    if not can_handle:
                        continue

                    suggestions.append({
                        'zona': tech_main_zone.get(t1, ''),
                        'subzona': o1['subzona'],
                        'origen': t1,
                        'destino': t2,
                        'order_id': o1['order_id'],
                        'franja': o1['franja'],
                        'address': o1.get('address', ''),
                        'distancia_estimada': f"{d_o1_t2:.2f} km (antes {d_o1_t1:.2f} km)",
                        'alerta': 'REASIGNACION PROXIMITY',
                        'pendientes_origen': tech_pending.get(t1, 0),
                        'pendientes_destino': tech_pending.get(t2, 0),
                        'justificacion': f"{t2} está {gain:.1f} km más cerca. Reasignar reduce desplazamiento sin afectar carga."
                    })

                    tech_movable[t1].remove(o1)
                    break

    # 6B: INTERCAMBIOS (SWAPS) - AHORA PERMITE FRANJAS DISTINTAS, respetando duplicada (salvo misma unidad)
    for t1 in all_techs_global:
        if not tech_movable.get(t1):
            continue
        for t2 in all_techs_global:
            if t2 <= t1 or not tech_movable.get(t2):
                continue

            z1 = tech_main_zone.get(t1, "SIN_ZONA").upper()
            z2 = tech_main_zone.get(t2, "SIN_ZONA").upper()

            if z1 != z2:
                continue

            c1 = get_tech_reference_point(t1, tech_all_orders, tech_locations)
            c2 = get_tech_reference_point(t2, tech_all_orders, tech_locations)

            orders_t1 = list(tech_movable.get(t1, []))
            orders_t2 = list(tech_movable.get(t2, []))

            for o1 in orders_t1:
                for o2 in orders_t2:
                    # ✅ Regla: INTERCAMBIOS solo en la MISMA franja
                    if o1.get('franja') != o2.get('franja'):
                        continue
                    # Sin coords y no misma unidad => no podemos justificar distancias
                    same_unit_flag = is_same_unit(o1, o2)
                    if (o1.get('lat', 0) == 0 or o2.get('lat', 0) == 0 or o1.get('lon', 0) == 0 or o2.get('lon', 0) == 0) and not same_unit_flag:
                        continue

                    # Validar que cada uno pueda recibir la orden del otro
                    can_t1, reason_t1 = can_tech_handle_franja(
                        tech_franja_counts.get(t1, {}),
                        tech_all_orders.get(t1, []),
                        o2['franja'],
                        current_hour,
                        allow_same_unit_override=True,
                        same_unit=same_unit_flag
                    )
                    if not can_t1:
                        continue

                    can_t2, reason_t2 = can_tech_handle_franja(
                        tech_franja_counts.get(t2, {}),
                        tech_all_orders.get(t2, []),
                        o1['franja'],
                        current_hour,
                        allow_same_unit_override=True,
                        same_unit=same_unit_flag
                    )
                    if not can_t2:
                        continue

                    # Distancias actuales vs swap (si hay coords)
                    if o1.get('lat', 0) and o2.get('lat', 0) and o1.get('lon', 0) and o2.get('lon', 0):
                        dist1_actual = haversine(o1['lat'], o1['lon'], c1[0], c1[1])
                        dist2_actual = haversine(o2['lat'], o2['lon'], c2[0], c2[1])

                        dist1_swap = haversine(o1['lat'], o1['lon'], c2[0], c2[1])
                        dist2_swap = haversine(o2['lat'], o2['lon'], c1[0], c1[1])

                        total_saved = (dist1_actual + dist2_actual) - (dist1_swap + dist2_swap)
                    else:
                        # sin coords pero misma unidad => no calculamos ahorro real
                        dist1_actual = dist2_actual = dist1_swap = dist2_swap = 0.0
                        total_saved = 0.0

                    is_same_zone = (z1 == z2)
                    threshold = MIN_SAVED_KM_FOR_SWAP if is_same_zone else 2.0

                    subzone_bonus = 0.0
                    if is_same_zone:
                        if (o1['subzona'] in tech_subzones.get(t2, set())) and (o2['subzona'] in tech_subzones.get(t1, set())):
                            subzone_bonus = 1.0

                    # Condición: ahorro suficiente O es misma unidad (para que torre decida)
                    allow_by_unit = same_unit_flag
                    if (total_saved + subzone_bonus) > threshold or allow_by_unit:
                        alert = "INTERCAMBIO"
                        if "EXCEPCION" in (reason_t1 + " " + reason_t2):
                            alert = "INTERCAMBIO (CONDICIONAL: MISMA UNIDAD)"
                        if allow_by_unit and (total_saved + subzone_bonus) <= threshold:
                            alert = "INTERCAMBIO (CONDICIONAL: MISMA UNIDAD)"

                        if o1.get('lat', 0) and o2.get('lat', 0):
                            dist_msg = f"Ahorro {total_saved:.1f} km" + (" + Subzona" if subzone_bonus > 0 else "")
                        else:
                            dist_msg = "MISMA UNIDAD (sin coords)"

                        suggestions.append({
                            'zona': z1 if is_same_zone else f"{z1} <-> {z2}",
                            'subzona': f"{o1['subzona']} <-> {o2['subzona']}",
                            'origen': t1,
                            'destino': t2,
                            'order_id': f"{o1['order_id']} -> {t2}  /  {o2['order_id']} -> {t1}",
                            'franja': o1.get('franja',''),
                            'address': f"{o1['order_id']} ({o1['franja']})  <->  {o2['order_id']} ({o2['franja']})",
                            'distancia_estimada': dist_msg,
                            'alerta': alert,
                            'pendientes_origen': tech_pending.get(t1, 0),
                            'pendientes_destino': tech_pending.get(t2, 0),
                            'justificacion': f"{dist_msg}. {reason_t1}; {reason_t2}. Torre de control valida si aplica."
                        })

                        # Remover para no repetir
                        if o1 in tech_movable.get(t1, []):
                            tech_movable[t1].remove(o1)
                        if o2 in tech_movable.get(t2, []):
                            tech_movable[t2].remove(o2)
                        break
                else:
                    continue
                break

    # =========================
    # EXPORT EXCEL (igual)
    # =========================
    wb_out = openpyxl.Workbook()

    ws_zona = wb_out.active
    ws_zona.title = "Resumen por Zona"
    zh = ['Zona', 'Tecnicos', 'Sin Asignar INI', 'Sin Asignar FIN', 'Pendientes INI', 'Pendientes FIN',
          'Finalizadas', 'Avg Ini', 'Desbalance Ini', 'Desbalance Fin']
    ws_zona.append(zh)
    style_header_row(ws_zona, len(zh))

    for s in zone_summaries:
        row_num = ws_zona.max_row + 1
        desb_ini = s['max_inicial'] - s['min_inicial']
        desb_fin = s.get('desbalance_final', 0)

        ws_zona.append([
            s['zona'], s['techs'], s['sin_asignar_inicial'], s['sin_asignar_final'],
            s['pendientes_inicial'], s['pendientes_final'],
            s['total_finalizadas'], s['avg_inicial'], desb_ini, desb_fin
        ])

        if desb_ini >= 3: ws_zona.cell(row=row_num, column=9).fill = ALERT_FILL
        elif desb_ini >= 2: ws_zona.cell(row=row_num, column=9).fill = WARN_FILL

        if desb_fin >= 3: ws_zona.cell(row=row_num, column=10).fill = ALERT_FILL
        elif desb_fin >= 2: ws_zona.cell(row=row_num, column=10).fill = WARN_FILL

    auto_fit_columns(ws_zona)

    ws_sub = wb_out.create_sheet("Distribucion por Tecnico")
    sh = ['Zona', 'Tecnico', 'Subzonas y Estados', 'Pend. Totales', 'Finalizadas', 'Carga Total', '# Subzonas']
    ws_sub.append(sh)
    style_header_row(ws_sub, len(sh))
    ws_sub.row_dimensions[1].height = 18

    subzone_summaries.sort(key=lambda x: (x['zona'], x['tecnico']))
    for s in subzone_summaries:
        row_num = ws_sub.max_row + 1
        ws_sub.append([
            s['zona'], s['tecnico'], s['subzonas_detalle'],
            s['pendientes'], s['finalizadas'], s['carga_total'], s['num_subzonas']
        ])
        ws_sub.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True, vertical='top')
        ws_sub.row_dimensions[row_num].height = max(18, s['num_subzonas'] * 16)

        if s['carga_total'] > MAX_ABSOLUTE_LOAD:
            ws_sub.cell(row=row_num, column=6).fill = ALERT_FILL
        elif s['carga_total'] >= MAX_IDEAL_LOAD:
            ws_sub.cell(row=row_num, column=6).fill = WARN_FILL
        elif s['pendientes'] <= 2:
            ws_sub.cell(row=row_num, column=4).fill = SUCCESS_FILL

    ws_sub.column_dimensions['C'].width = 60
    ws_sub.column_dimensions['A'].width = 16
    ws_sub.column_dimensions['B'].width = 28
    for col_letter in ['D', 'E', 'F', 'G']:
        ws_sub.column_dimensions[col_letter].width = 14

    ws_sug = wb_out.create_sheet("Sugerencias de Movimientos")
    sugh = [
        'Zona', 'Subzona', 'Direccion', 'De Tecnico (Origen)',
        'Pendientes Origen', 'A Tecnico (Destino)', 'Pendientes Destino',
        'ID Orden', 'Franja', 'Distancia Aprox.', 'Alerta', 'Justificación'
    ]
    ws_sug.append(sugh)
    style_header_row(ws_sug, len(sugh))

    for sug in suggestions:
        row_num = ws_sug.max_row + 1
        addr = sug.get('address', '') or sug.get('subzona', '')
        ws_sug.append([
            sug.get('zona', ''), sug.get('subzona', ''), addr,
            sug.get('origen', ''), sug.get('pendientes_origen', ''),
            sug.get('destino', ''), sug.get('pendientes_destino', ''),
            sug.get('order_id', ''), sug.get('franja', ''),
            sug.get('distancia_estimada', ''), sug.get('alerta', ''),
            sug.get('justificacion', '')
        ])
        if sug.get('alerta'):
            ws_sug.cell(row=row_num, column=11).fill = WARN_FILL

    auto_fit_columns(ws_sug)


    ws_disc = wb_out.create_sheet("Movimientos Descartados")
    dh = ['Zona', 'Subzona', 'Direccion', 'ID Orden', 'Origen', 'Destino Evaluado', 'Franja', 'Motivo', 'Detalle']
    ws_disc.append(dh)
    style_header_row(ws_disc, len(dh))
    for d in discarded_suggestions:
        ws_disc.append([
            d.get('zona',''), d.get('subzona',''), d.get('direccion',''), d.get('order_id',''),
            d.get('origen',''), d.get('destino',''), d.get('franja',''), d.get('motivo',''), d.get('detalle','')
        ])
    auto_fit_columns(ws_disc)

    ws_alert = wb_out.create_sheet("Alertas")
    ah = ['Tipo', 'Zona', 'Tecnico', 'Detalle']
    ws_alert.append(ah)
    style_header_row(ws_alert, len(ah))

    priority_map = {'SOBRECARGA': 0, 'FRANJA EN RIESGO': 1, 'BAJO AVANCE RELATIVO': 2, 'FRANJAS DUPLICADAS': 3,
                    'EXCESO TARDE': 4, 'FRANJA AJUSTADA': 5, 'MULTI-ESTADO ACTIVO': 6}
    alerts.sort(key=lambda a: priority_map.get(a['tipo'], 99))

    if alerts:
        for a in alerts:
            row_num = ws_alert.max_row + 1
            ws_alert.append([a['tipo'], a['zona'], a['tecnico'], a['detalle']])
            if a['tipo'] in ['SOBRECARGA', 'FRANJA EN RIESGO']:
                ws_alert.cell(row=row_num, column=1).fill = ALERT_FILL
            elif a['tipo'] in ['BAJO AVANCE RELATIVO', 'FRANJAS DUPLICADAS', 'EXCESO TARDE', 'MULTI-ESTADO ACTIVO']:
                ws_alert.cell(row=row_num, column=1).fill = WARN_FILL
    else:
        ws_alert.append(['OK', 'N/A', 'N/A', 'No se detectaron alertas'])

    auto_fit_columns(ws_alert)

    try:
        date_tag = now_bogota().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"sugerencias_nivelacion_{date_tag}.xlsx"
        output_path = os.path.join(os.getcwd(), filename)
        wb_out.save(output_path)

        msg_parts = [
            "Nivelacion completada.",
            f"  Sugerencias generadas: {len(suggestions)}",
            f"  Alertas detectadas: {len(alerts)}",
            f"  Movimientos descartados: {len(discarded_suggestions)}",
            f"  Zonas analizadas: {len(zones)}",
            f"  Total ordenes: {len(data)}",
        ]
        if not suggestions:
            msg_parts.append("\n  NOTA: No se generaron movimientos.")
            msg_parts.append("  La carga esta balanceada o las restricciones impiden mover.")

        return "\n".join(msg_parts), output_path
    except Exception as e:
        return f"Error guardando reporte: {str(e)}", None        return f"Error guardando reporte: {str(e)}", None
