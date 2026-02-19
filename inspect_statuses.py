import glob
import os
import openpyxl

def inspect_statuses():
    files = glob.glob("*.xlsx")
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
         print("No files found.")
         return
    files.sort(key=os.path.getmtime, reverse=True)
    latest = files[0]
    print(f"Reading: {latest}")
    
    wb = openpyxl.load_workbook(latest, read_only=True, data_only=True)
    sheet = wb.active
    
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    
    status_col_names = ['status_txt', 'Estado', 'Estado de orden de trabajo']
    idx = -1
    for h in headers:
        if str(h).strip() in status_col_names:
            idx = headers.index(h)
            break
            
    if idx == -1:
        print("Status column not found in headers:", headers)
        return

    statuses = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        val = row[idx]
        if val:
            statuses.add(str(val).strip())
            
    print("\nUnique Statuses found:")
    for s in statuses:
        print(f"'{s}'")

inspect_statuses()
