import openpyxl
import glob
import os

def get_latest_preruta_file():
    files = glob.glob("*pre_ruta*.xlsx")
    if not files: return None
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]

def inspect():
    f = get_latest_preruta_file()
    if not f:
        print("No file found.")
        return

    print(f"Inspecting: {f}")
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    sheet = wb.active
    
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    idx_status = -1
    for i, h in enumerate(headers):
        if str(h).strip() == 'Estado': idx_status = i
            
    statuses = set()
    empty_count = 0
    total_rows = 0
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        val = row[idx_status]
        if val:
            statuses.add(str(val).strip())
        else:
            statuses.add("EMPTY/NONE")
            empty_count += 1
            
    print(f"\nTotal Rows: {total_rows}")
    print(f"Empty Status Rows: {empty_count}")
    print("\n--- Unique Statuses Found ---")
    for s in sorted(list(statuses)):
        print(f"'{s}'")

if __name__ == "__main__":
    inspect()
