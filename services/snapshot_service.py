"""
services/snapshot_service.py
Reporte diario de seguimiento fotografico.

Reglas:
- Cada foto/corte guarda fecha y hora exacta.
- Se permiten varios cortes durante el mismo dia.
- Solo existe reporte descargable del dia actual.
- Al cambiar el dia, los cortes anteriores vencen y no se descargan.
- El seguimiento principal se hace por franja horaria y tipo de orden.
- Si una franja/tipo baja de cantidad y no hay canceladas que expliquen la baja,
  la diferencia se marca como reprogramada para el resumen operativo.
- La lista de ordenes reprogramadas solo incluye ordenes que desaparecen del corte y no quedan explicadas por canceladas.
"""
import io
import json
import logging
import os
import tempfile
from datetime import datetime

logger = logging.getLogger(__name__)

_store: dict = {}
_STORE_FILE = os.environ.get(
    "REPORT_SNAPSHOT_FILE",
    os.path.join(tempfile.gettempdir(), "nivelacion_pro_reporte_hoy.json"),
)

_GRUPOS = {
    "Programadas": {"programado", "programada"},
    "Por Programar": {"por programar"},
    "En Proceso": {
        "en camino", "en sitio", "iniciado", "iniciada",
        "mac enviada", "mac principal enviada",
        "dispositivos subidos", "dispositivos cargados",
    },
    "Finalizadas": {"finalizado", "finalizada", "cerrado", "cerrada", "completado", "completada"},
    "Por Auditar": {"por auditar"},
    "Canceladas": {"cancelado", "cancelada", "cancelada por cliente", "cancelado por cliente"},
}


def _now_naive() -> datetime:
    from config import now_bogota
    dt = now_bogota()
    return dt.replace(tzinfo=None) if getattr(dt, "tzinfo", None) else dt


def _today() -> str:
    return _now_naive().strftime("%Y-%m-%d")


def _strip_accents(text: str) -> str:
    value = str(text or "").strip().lower()
    for a, b in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n")):
        value = value.replace(a, b)
    return " ".join(value.split())


def _norm_estado(value) -> str:
    try:
        from services.normalization import norm_status
        return norm_status(value)
    except Exception:
        return _strip_accents(value)


def _estado_grupo(estado) -> str:
    s = _norm_estado(estado)
    if "cancelad" in s:
        return "Canceladas"
    if "por auditar" in s:
        return "Por Auditar"
    if any(x in s for x in ("finaliz", "cerrad", "completad")):
        return "Finalizadas"
    if "por programar" in s:
        return "Por Programar"
    if "programad" in s or "programado" in s:
        return "Programadas"
    if any(x in s for x in ("en camino", "en sitio", "iniciad", "mac", "dispositivos")):
        return "En Proceso"
    return "Otros"


def _norm_franja(value) -> str:
    franja = str(value or "").strip()
    if not franja or franja.lower() in ("none", "nan", "sin valor"):
        return "Sin Franja"
    return franja


def _norm_tipo(value) -> str:
    """Agrupa el tipo de trabajo SOLO para el informe diario.

    Para que el Excel no crezca demasiado, el reporte usa tres grupos:
    - Instalacion
    - Soporte
    - Otros Tipos

    Esta funcion no cambia la data original ni la logica del dashboard operativo;
    solo simplifica el agrupamiento del seguimiento fotografico.
    """
    raw = _strip_accents(value)
    if not raw or raw in ("none", "nan", "sin valor"):
        return "Otros Tipos"
    if "soporte" in raw:
        return "Soporte"
    if "instal" in raw:
        return "Instalacion"
    return "Otros Tipos"




def _merge_key_stats(dest: dict, src: dict) -> dict:
    """Suma estadisticas de una franja/tipo al grupo simplificado."""
    dest.setdefault("total", 0)
    dest.setdefault("vigentes", 0)
    dest.setdefault("canceladas", 0)
    dest.setdefault("por_estado", {})
    src = src or {}
    dest["total"] += int(src.get("total", 0) or 0)
    dest["vigentes"] += int(src.get("vigentes", 0) or 0)
    dest["canceladas"] += int(src.get("canceladas", 0) or 0)
    for estado, qty in (src.get("por_estado") or {}).items():
        dest["por_estado"][estado] = dest["por_estado"].get(estado, 0) + int(qty or 0)
    return dest


def _normalizar_snapshot_tipos(corte: dict) -> dict:
    """Colapsa tipos antiguos del reporte a Instalacion, Soporte y Otros Tipos.

    Esto es importante porque pueden existir fotos tomadas antes del ajuste,
    por ejemplo Add: Extension, Add: Router o Traslado. El informe debe
    mostrarlas agrupadas sin exigir borrar el corte del dia.
    """
    if not isinstance(corte, dict):
        return corte

    # por_tipo
    por_tipo = {}
    for tipo, qty in (corte.get("por_tipo") or {}).items():
        grupo = _norm_tipo(tipo)
        por_tipo[grupo] = por_tipo.get(grupo, 0) + int(qty or 0)
    corte["por_tipo"] = dict(sorted(por_tipo.items()))

    # por_franja_tipo
    nuevo_ft = {}
    for franja, tipos in (corte.get("por_franja_tipo") or {}).items():
        franja_norm = _norm_franja(franja)
        fr_data = nuevo_ft.setdefault(franja_norm, {})
        for tipo, stats in (tipos or {}).items():
            grupo = _norm_tipo(tipo)
            _merge_key_stats(fr_data.setdefault(grupo, _empty_key_stats()), stats or {})
    corte["por_franja_tipo"] = {f: dict(sorted(t.items())) for f, t in sorted(nuevo_ft.items())}

    # order_state interno para comparaciones entre fotos.
    if "_order_state" in corte:
        for _oid, state in (corte.get("_order_state") or {}).items():
            if not isinstance(state, dict):
                continue
            franja = _norm_franja(state.get("franja"))
            tipo = _norm_tipo(state.get("tipo"))
            state["franja"] = franja
            state["tipo"] = tipo
            state["key"] = f"{franja}||{tipo}"

    # detalle y diferencias que vengan de cortes viejos tambien deben verse agrupados.
    for d in corte.get("detalle_cambios", []) or []:
        if isinstance(d, dict) and "tipo_orden" in d:
            d["tipo_orden"] = _norm_tipo(d.get("tipo_orden"))

    return corte


def _recalcular_diferencias_dia(cortes: list) -> list:
    """Recalcula diferencias con tipos agrupados para evitar filas duplicadas viejas."""
    normalizados = [_normalizar_snapshot_tipos(c) for c in (cortes or []) if isinstance(c, dict)]
    for idx, corte in enumerate(normalizados):
        corte["reprogramadas"] = 0
        corte["nuevas"] = 0
        corte["salieron"] = 0
        corte["cambios_estado"] = 0
        corte["cambios_franja"] = 0
        corte["cambios_tipo"] = 0
        corte["canceladas_nuevas"] = 0
        corte["detalle_cambios"] = []
        corte["diferencias_franja_tipo"] = []
        corte["ordenes_reprogramadas"] = []
        if idx > 0:
            _comparar(normalizados[idx - 1], corte)
    return normalizados



def _normalizar_hora_manual(value) -> str:
    """Convierte una hora indicada por el usuario a HH:MM:SS.

    Acepta valores como 8, 8:00, 08:00, 13:30 o 17:00:00.
    Si no es valida, devuelve cadena vacia y se usa la hora real de captura.
    """
    raw = str(value or "").strip()
    if not raw:
        return ""
    raw = raw.replace(".", ":")
    parts = raw.split(":")
    try:
        if len(parts) == 1:
            h, m, sec = int(parts[0]), 0, 0
        elif len(parts) == 2:
            h, m, sec = int(parts[0]), int(parts[1]), 0
        else:
            h, m, sec = int(parts[0]), int(parts[1]), int(parts[2])
        if not (0 <= h <= 23 and 0 <= m <= 59 and 0 <= sec <= 59):
            return ""
        return f"{h:02d}:{m:02d}:{sec:02d}"
    except Exception:
        return ""

def _hora_label(dt: datetime) -> str:
    h = dt.hour
    if 5 <= h < 10:
        return f"Manana {dt.strftime('%H:%M:%S')}"
    if 10 <= h < 14:
        return f"Mediodia {dt.strftime('%H:%M:%S')}"
    if 14 <= h < 19:
        return f"Tarde {dt.strftime('%H:%M:%S')}"
    return f"Cierre {dt.strftime('%H:%M:%S')}"


def _purge_old_days() -> None:
    global _store
    hoy = _today()
    if _store and set(_store.keys()) != {hoy}:
        _store = {hoy: _store.get(hoy, [])}
    try:
        if os.path.exists(_STORE_FILE):
            with open(_STORE_FILE, "r", encoding="utf-8") as fh:
                payload = json.load(fh)
            if payload.get("fecha") != hoy:
                os.remove(_STORE_FILE)
    except Exception as exc:
        logger.warning("No se pudo depurar reporte diario: %s", exc)


def _load_store() -> None:
    global _store
    hoy = _today()
    _purge_old_days()
    if _store.get(hoy):
        return
    try:
        if not os.path.exists(_STORE_FILE):
            _store.setdefault(hoy, [])
            return
        with open(_STORE_FILE, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
        cortes = payload.get("cortes", []) or [] if payload.get("fecha") == hoy else []
        _store = {hoy: _recalcular_diferencias_dia(cortes)}
        _save_store()
    except Exception as exc:
        logger.warning("No se pudo cargar reporte diario: %s", exc)
        _store.setdefault(hoy, [])


def _save_store() -> None:
    hoy = _today()
    os.makedirs(os.path.dirname(_STORE_FILE), exist_ok=True)
    tmp = f"{_STORE_FILE}.tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump({"fecha": hoy, "cortes": _store.get(hoy, [])}, fh, ensure_ascii=False)
    os.replace(tmp, _STORE_FILE)


def _order_id(order: dict, pos: int) -> str:
    oid = str(order.get("id") or order.get("orden") or order.get("appointment_id") or "").strip()
    return oid if oid and oid.lower() not in ("none", "nan") else f"row_{pos}"


def _empty_key_stats() -> dict:
    return {
        "total": 0,
        "vigentes": 0,
        "canceladas": 0,
        "por_estado": {},
    }


def _clasificar(orders: list) -> dict:
    por_estado = {g: 0 for g in _GRUPOS}
    por_estado.setdefault("Otros", 0)
    por_franja: dict = {}
    por_tipo: dict = {}
    por_franja_tipo: dict = {}
    order_state: dict = {}

    for pos, o in enumerate(orders or []):
        if not isinstance(o, dict):
            continue
        oid = _order_id(o, pos)
        estado_raw = o.get("estado", "")
        estado_norm = _norm_estado(estado_raw)
        grupo = _estado_grupo(estado_raw)
        franja = _norm_franja(o.get("franja", "Sin Franja"))
        tipo = _norm_tipo(o.get("tipo", "Otro"))
        tecnico = str(o.get("tecnico", "SIN_ASIGNAR") or "SIN_ASIGNAR").strip() or "SIN_ASIGNAR"

        por_estado[grupo] = por_estado.get(grupo, 0) + 1
        por_franja[franja] = por_franja.get(franja, 0) + 1
        por_tipo[tipo] = por_tipo.get(tipo, 0) + 1

        fr_data = por_franja_tipo.setdefault(franja, {})
        key_data = fr_data.setdefault(tipo, _empty_key_stats())
        key_data["total"] += 1
        key_data["por_estado"][grupo] = key_data["por_estado"].get(grupo, 0) + 1
        if grupo == "Canceladas":
            key_data["canceladas"] += 1
        else:
            key_data["vigentes"] += 1

        order_state[oid] = {
            "estado": estado_norm,
            "grupo": grupo,
            "franja": franja,
            "tipo": tipo,
            "key": f"{franja}||{tipo}",
            "tecnico": tecnico,
        }

    return {
        "total": len(order_state),
        "vigentes": sum(1 for o in order_state.values() if o.get("grupo") != "Canceladas"),
        "por_estado": por_estado,
        "por_franja": dict(sorted(por_franja.items())),
        "por_tipo": dict(sorted(por_tipo.items())),
        "por_franja_tipo": {f: dict(sorted(t.items())) for f, t in sorted(por_franja_tipo.items())},
        "_order_state": order_state,
        "reprogramadas": 0,
        "nuevas": 0,
        "salieron": 0,
        "cambios_estado": 0,
        "cambios_franja": 0,
        "cambios_tipo": 0,
        "canceladas_nuevas": 0,
        "detalle_cambios": [],
        "diferencias_franja_tipo": [],
        "ordenes_reprogramadas": [],
    }


def _flatten_franja_tipo(corte: dict) -> dict:
    out = {}
    for franja, tipos in (corte.get("por_franja_tipo") or {}).items():
        for tipo, data in (tipos or {}).items():
            out[f"{franja}||{tipo}"] = {
                "franja": franja,
                "tipo": tipo,
                "total": int((data or {}).get("total", 0) or 0),
                "vigentes": int((data or {}).get("vigentes", 0) or 0),
                "canceladas": int((data or {}).get("canceladas", 0) or 0),
            }
    return out


def _total_canceladas(corte: dict) -> int:
    return int(((corte.get("por_estado") or {}).get("Canceladas", 0)) or 0)


def _comparar(anterior: dict, actual: dict) -> None:
    """Compara dos cortes consecutivos para el informe operativo.

    Regla actual del negocio:
    - "Reprogramada" no es una modificacion de estado/tipo ni un cambio interno.
    - Se cuenta como reprogramacion operativa cuando el appointment estaba en el corte anterior
      y ya no aparece en el siguiente corte.
    - Si la baja esta explicada por nuevas canceladas del corte, no se cuenta como reprogramada.
    - La lista solo muestra numeros de orden que salieron y no fueron explicados por canceladas.
    """
    prev = anterior.get("_order_state", {}) or {}
    curr = actual.get("_order_state", {}) or {}
    prev_ids = set(prev.keys())
    curr_ids = set(curr.keys())

    salieron = sorted(prev_ids - curr_ids)
    nuevas = sorted(curr_ids - prev_ids)
    comunes = sorted(prev_ids & curr_ids)

    cambios_estado = 0
    cambios_franja = 0
    cambios_tipo = 0
    common_became_cancelled = 0

    for oid in comunes:
        p = prev.get(oid) or {}
        c = curr.get(oid) or {}
        if p.get("grupo") != c.get("grupo"):
            cambios_estado += 1
        if p.get("franja") != c.get("franja"):
            cambios_franja += 1
        if p.get("tipo") != c.get("tipo"):
            cambios_tipo += 1
        if c.get("grupo") == "Canceladas" and p.get("grupo") != "Canceladas":
            common_became_cancelled += 1

    prev_vigentes = int(anterior.get("vigentes", anterior.get("total", 0)) or 0)
    curr_vigentes = int(actual.get("vigentes", actual.get("total", 0)) or 0)
    bajaron_vigentes = max(prev_vigentes - curr_vigentes, 0)
    subieron_vigentes = max(curr_vigentes - prev_vigentes, 0)

    cancel_delta_total = max(_total_canceladas(actual) - _total_canceladas(anterior), 0)
    canceladas_nuevas = max(cancel_delta_total, common_became_cancelled)

    # Cantidad operativa de reprogramadas: baja de appointments vigentes que no se explica por canceladas.
    reprogramadas_total = max(bajaron_vigentes - canceladas_nuevas, 0)

    # Lista de ordenes: solo ordenes que desaparecieron del corte. No se listan modificaciones.
    # Si hubo canceladas nuevas, se usan para explicar primero parte de las salidas.
    salieron_ordenados = sorted(
        salieron,
        key=lambda oid: (
            (prev.get(oid) or {}).get("franja", "Sin Franja") == "Sin Franja",
            (prev.get(oid) or {}).get("franja", "Sin Franja"),
            str(oid),
        ),
    )
    omitidas_por_cancelacion = min(len(salieron_ordenados), canceladas_nuevas)
    candidatos_reprogramados = salieron_ordenados[omitidas_por_cancelacion:]
    if reprogramadas_total:
        candidatos_reprogramados = candidatos_reprogramados[:reprogramadas_total]
    else:
        candidatos_reprogramados = []

    ordenes_reprogramadas = []
    for oid in candidatos_reprogramados:
        p = prev.get(oid) or {}
        ordenes_reprogramadas.append({
            "orden": str(oid),
            "franja": p.get("franja") or "Sin Franja",
            "franja_antes": p.get("franja") or "Sin Franja",
            "franja_despues": "No aparece en el corte",
            "tipo": p.get("tipo") or "",
        })

    actual["salieron"] = len(salieron)
    actual["nuevas"] = subieron_vigentes if subieron_vigentes else len(nuevas)
    actual["reprogramadas"] = reprogramadas_total
    actual["cambios_estado"] = cambios_estado
    actual["cambios_franja"] = cambios_franja
    actual["cambios_tipo"] = cambios_tipo
    actual["canceladas_nuevas"] = canceladas_nuevas
    actual["detalle_cambios"] = []
    actual["ordenes_reprogramadas"] = ordenes_reprogramadas[:500]
    actual["diferencias_franja_tipo"] = []


def registrar_corte(orders: list, label: str = None, hora_manual: str = None) -> dict:
    _load_store()
    now = _now_naive()
    fecha = now.strftime("%Y-%m-%d")
    hora_exacta = now.strftime("%H:%M:%S")
    hora_operativa = _normalizar_hora_manual(hora_manual) or hora_exacta
    etiqueta = label or f"Corte {hora_operativa}"

    stats = _clasificar(orders)
    stats["fecha"] = fecha
    stats["hora"] = hora_exacta
    stats["hora_exacta"] = hora_exacta
    stats["label"] = etiqueta
    dia = _store.setdefault(fecha, [])
    if dia:
        # Por compatibilidad, normaliza cortes anteriores que hayan quedado con tipos crudos
        # como Add: Extension, Add: Router, Traslado, etc.
        dia[:] = _recalcular_diferencias_dia(dia)
        _comparar(dia[-1], stats)

    corte = {
        "id": f"{fecha}_{hora_exacta}_{len(dia) + 1}",
        "label": etiqueta,
        "fecha": fecha,
        "hora": hora_operativa,
        "hora_exacta": hora_exacta,
        "hora_captura": hora_exacta,
        "hora_manual": hora_operativa if hora_operativa != hora_exacta else "",
        "timestamp": now.isoformat(sep=" ", timespec="seconds"),
        "total": stats["total"],
        "vigentes": stats["vigentes"],
        "por_estado": stats["por_estado"],
        "por_franja": stats["por_franja"],
        "por_tipo": stats["por_tipo"],
        "por_franja_tipo": stats["por_franja_tipo"],
        "reprogramadas": stats["reprogramadas"],
        "salieron": stats["salieron"],
        "nuevas": stats["nuevas"],
        "cambios_estado": stats["cambios_estado"],
        "cambios_franja": stats["cambios_franja"],
        "cambios_tipo": stats["cambios_tipo"],
        "canceladas_nuevas": stats["canceladas_nuevas"],
        "detalle_cambios": stats["detalle_cambios"],
        "ordenes_reprogramadas": stats.get("ordenes_reprogramadas", []),
        "diferencias_franja_tipo": stats["diferencias_franja_tipo"],
        "_order_state": stats["_order_state"],
    }
    dia.append(corte)
    _save_store()
    logger.info(
        "Corte %s %s | total=%s vigentes=%s cancel=%s reprog=%s",
        fecha, hora_exacta, corte["total"], corte["vigentes"],
        corte["canceladas_nuevas"], corte["reprogramadas"],
    )
    return corte


def _public(corte: dict) -> dict:
    return {k: v for k, v in corte.items() if not k.startswith("_")}


def get_cortes(fecha: str = None) -> list:
    _load_store()
    hoy = _today()
    if fecha and fecha != hoy:
        return []
    cortes = _recalcular_diferencias_dia(_store.get(hoy, []))
    _store[hoy] = cortes
    return [_public(c) for c in cortes]


def get_fechas() -> list:
    _load_store()
    hoy = _today()
    return [hoy] if _store.get(hoy) else []


def _all_franja_tipo_keys(cortes: list) -> list:
    keys = set()
    for c in cortes:
        keys.update(_flatten_franja_tipo(c).keys())
    def sort_key(key):
        fr, tp = key.split("||", 1)
        return (fr == "Sin Franja", fr, tp)
    return sorted(keys, key=sort_key)



def get_resumen_ejecutivo(fecha: str = None) -> dict:
    """Resumen liviano del informe diario.

    Usa el primer y ultimo corte del dia actual. Las canceladas y reprogramadas
    son la suma detectada entre cortes.
    """
    cortes = get_cortes(fecha)
    if not cortes:
        return {
            "appointments_inicio": 0,
            "appointments_final": 0,
            "cancelados": 0,
            "reprogramados": 0,
            "hora_inicio": "",
            "hora_final": "",
        }
    primero = cortes[0]
    ultimo = cortes[-1]
    return {
        "appointments_inicio": int(primero.get("vigentes", primero.get("total", 0)) or 0),
        "appointments_final": int(ultimo.get("vigentes", ultimo.get("total", 0)) or 0),
        "cancelados": sum(int(c.get("canceladas_nuevas", 0) or 0) for c in cortes),
        "reprogramados": sum(int(c.get("reprogramadas", 0) or 0) for c in cortes),
        "hora_inicio": primero.get("hora", ""),
        "hora_final": ultimo.get("hora", ""),
    }


def get_ordenes_reprogramadas_consolidadas(fecha: str = None) -> list:
    """Unica lista consolidada de ordenes realmente reprogramadas, por franja.

    Solo incluye ordenes que estaban en el corte anterior y ya no aparecen en el siguiente,
    descontando las bajas explicadas por canceladas. No incluye modificaciones de estado/tipo.
    """
    cortes = get_cortes(fecha)
    out = []
    seen = set()
    for c in cortes:
        for item in c.get("ordenes_reprogramadas", []) or []:
            orden = str(item.get("orden", "")).strip()
            if not orden or orden in seen:
                continue
            seen.add(orden)
            out.append({
                "hora": c.get("hora", ""),
                "orden": orden,
                "franja_antes": item.get("franja_antes", ""),
                "franja_despues": item.get("franja_despues", "No aparece en el corte"),
                "franja": item.get("franja", item.get("franja_antes", "Sin Franja")),
            })
    def _key(x):
        fr = x.get("franja") or "Sin Franja"
        return (fr == "Sin Franja", fr, x.get("hora", ""), x.get("orden", ""))
    return sorted(out, key=_key)


def reset_reporte_diario() -> bool:
    """Borra manualmente los cortes del dia actual para iniciar limpio el informe."""
    _load_store()
    hoy = _today()
    _store[hoy] = []
    try:
        if os.path.exists(_STORE_FILE):
            os.remove(_STORE_FILE)
    except Exception as exc:
        logger.warning("No se pudo eliminar archivo de reporte diario: %s", exc)
    return True


def generar_excel(fecha: str = None) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl no instalado")

    hoy = _today()
    if fecha and fecha != hoy:
        raise ValueError("El reporte solicitado ya vencio. Solo se puede descargar el informe del dia actual.")
    fecha = hoy
    cortes = get_cortes(fecha)
    resumen = get_resumen_ejecutivo(fecha)
    reprogramadas = get_ordenes_reprogramadas_consolidadas(fecha)

    wb = Workbook()

    def font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size)

    def fill(color):
        return PatternFill("solid", fgColor=color)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    hdr = "1F3864"
    title = "D6E4F0"

    def style_header(row):
        for cell in row:
            cell.font = font(True, "FFFFFF")
            cell.fill = fill(hdr)
            cell.alignment = center
            cell.border = thin

    def autosize(ws, max_width=32):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = min(max_width, max(10, max(len(str(c.value or "")) for c in col) + 2))
            ws.column_dimensions[letter].width = width

    ws = wb.active
    ws.title = "Resumen Ejecutivo"
    ws.append([f"Reporte diario de seguimiento operativo - {fecha}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws["A1"].font = font(True, hdr, 13)
    ws["A1"].fill = fill(title)
    ws["A1"].alignment = center
    ws.append([])
    ws.append(["Concepto", "Valor"])
    style_header(ws[3])
    rows = [
        ("Hora primer corte", resumen.get("hora_inicio", "")),
        ("Hora ultimo corte", resumen.get("hora_final", "")),
        ("Appointments inicio del corte", resumen.get("appointments_inicio", 0)),
        ("Appointments final del corte", resumen.get("appointments_final", 0)),
        ("Total appointments cancelados", resumen.get("cancelados", 0)),
        ("Total appointments reprogramados", resumen.get("reprogramados", 0)),
    ]
    for row in rows:
        ws.append(list(row))
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border = thin
            cell.alignment = left if cell.column == 1 else center
    autosize(ws, 45)
    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("Cortes del Dia")
    ws2.append(["Hora corte", "Hora captura real", "Appointments vigentes", "Instalacion", "Soporte", "Otros Tipos", "Canceladas acumuladas", "Reprogramados del corte"])
    style_header(ws2[1])
    for c in cortes:
        por_tipo = c.get("por_tipo") or {}
        ws2.append([
            c.get("hora"), c.get("hora_captura") or c.get("hora_exacta") or c.get("hora"),
            c.get("vigentes", c.get("total", 0)),
            por_tipo.get("Instalacion", 0), por_tipo.get("Soporte", 0), por_tipo.get("Otros Tipos", 0),
            (c.get("por_estado") or {}).get("Canceladas", 0), c.get("reprogramadas", 0),
        ])
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin
            cell.alignment = center
    autosize(ws2)
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Reprogramadas")
    ws3.append(["Franja", "Orden", "Observacion", "Hora detectada"])
    style_header(ws3[1])
    for item in reprogramadas:
        ws3.append([item.get("franja"), item.get("orden"), item.get("franja_despues") or "No aparece en el corte", item.get("hora")])
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin
            cell.alignment = left if cell.column in (1, 2, 3) else center
    autosize(ws3)
    ws3.freeze_panes = "A2"

    ws4 = wb.create_sheet("Notas")
    ws4.append(["Regla aplicada"])
    ws4.append(["El reporte solo corresponde al dia actual. Al cambiar el dia, el reporte anterior vence y no queda disponible para descarga."])
    ws4.append(["El resumen ejecutivo compara el primer corte contra el ultimo corte disponible del dia."])
    ws4.append(["Las canceladas se calculan entre cortes. Si baja la cantidad de appointments vigentes y la baja no esta explicada por canceladas, se toma como reprogramacion operativa."])
    ws4.append(["La lista Reprogramadas solo muestra ordenes que estaban en el corte anterior y no aparecen en el siguiente corte, despues de descontar canceladas. No incluye modificaciones de estado/tipo."])
    ws4.append(["Para reducir el tamaño, los tipos se agrupan en Instalacion, Soporte y Otros Tipos."])
    autosize(ws4, 90)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
