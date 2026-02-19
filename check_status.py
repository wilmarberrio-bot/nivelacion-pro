import openpyxl

INPUT_FILE = r"c:\Users\Usuario\OneDrive\Documentos\nivelacion\pre_ruta_2_0_2026-02-17T10_51_45.802611833-05_00.xlsx"

try:
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
    sheet = wb.active
    
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    print(f"Headers: {headers}")
    
    # Find Estado column
    idx_estado = -1
    for i, h in enumerate(headers):
        if str(h).strip() == 'Estado':
            idx_estado = i
            break
            
    if idx_estado != -1:
        unique_statuses = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[idx_estado]:
                unique_statuses.add(str(row[idx_estado]).strip())
        print(f"Unique Statuses: {unique_statuses}")
    else:
        print("Column 'Estado' not found.")

except Exception as e:
    print(e)
