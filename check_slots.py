import openpyxl
import os
import glob

def get_latest_preruta_file():
    files = glob.glob("*.xlsx")
    candidates = [f for f in files
                  if ("pre_ruta" in f.lower() or "nivelacion" in f.lower())
                  and not f.startswith("~$") and not f.startswith("sugerencias_")]
    if not candidates:
        return None
    candidates.sort(key=os.path.getmtime, reverse=True)
    return candidates[0]

def check_slots(filename):
    wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
    sheet = wb.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    
    idx_franja = -1
    for i, h in enumerate(headers):
        if h and str(h).lower() in ['franja_label', 'franja', 'cita', 'ventana']:
            idx_franja = i
            break
            
    if idx_franja == -1:
        print("No franja column found")
        return

    franja_counts = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        f = str(row[idx_franja]) if row[idx_franja] else "None"
        franja_counts[f] = franja_counts.get(f, 0) + 1
        
    print(f"File: {filename}")
    for f, c in sorted(franja_counts.items()):
        print(f"{f}: {c}")

if __name__ == '__main__':
    f = get_latest_preruta_file()
    if f:
        check_slots(f)
    else:
        print("No file found")
