import openpyxl
from datetime import datetime
import os

# --- CONFIGURATION ---
INPUT_FILE = r"c:\Users\Usuario\OneDrive\Documentos\nivelacion\pre_ruta_2_0_2026-02-17T10_51_45.802611833-05_00.xlsx"
DATE_STR = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
OUTPUT_FILE = fr"c:\Users\Usuario\OneDrive\Documentos\nivelacion\reporte_nivelacion_{DATE_STR}.xlsx"

# Status Mapping
STATUS_COLOR = {
    'Finalizada': '00FF00', # Green (Available?)
    'En sitio': 'FFFF00', # Yellow (Working)
    'Iniciada': 'FFFF00', # Yellow
    'Programada': 'CCCCCC', # Grey
}

def get_col_index(headers, possible_names):
    if isinstance(possible_names, str):
        possible_names = [possible_names]
    for i, h in enumerate(headers):
        if h and str(h).strip().lower() in [n.lower() for n in possible_names]:
            return i
    return -1

def generate_report():
    print(f"Loading data from {INPUT_FILE}...")
    try:
        wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
        sheet = wb.active
    except Exception as e:
        print(f"Error loading file: {e}")
        return

    # Get headers
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    # Map columns
    idx_status = get_col_index(headers, ['Estado', 'Estado de orden de trabajo'])
    idx_type = get_col_index(headers, 'Tipo de orden')
    idx_zone = get_col_index(headers, ['Zone Name', 'Zona', 'zona_op'])
    idx_subzone = get_col_index(headers, 'Subzone')
    idx_tech = get_col_index(headers, ['technenvician', 'Técnico asignado', 'Técnico'])
    idx_order = get_col_index(headers, ['ID', 'Número de orden'])
    
    missing = []
    if idx_status == -1: missing.append("Estado")
    if idx_type == -1: missing.append("Tipo de orden")
    if idx_zone == -1: missing.append("Zone")
    if idx_subzone == -1: missing.append("Subzone")
    if idx_tech == -1: missing.append("Technician")
    if idx_order == -1: missing.append("Order ID")

    if missing:
        print(f"Error: Missing required columns: {missing}")
        return

    print("Processing data...")
    
    tech_status = {} # Technician -> Latest Status Information
    tech_orders = [] # List of all tech orders

    for row in sheet.iter_rows(min_row=2, values_only=True):
        status = str(row[idx_status]).strip() if row[idx_status] else ""
        order_type = str(row[idx_type]).strip() if row[idx_type] else ""
        zona = str(row[idx_zone]).strip() if row[idx_zone] else ""
        subzona = str(row[idx_subzone]).strip() if row[idx_subzone] else ""
        tech = str(row[idx_tech]).strip() if row[idx_tech] else ""
        order_num = row[idx_order]

        if tech:
            tech_orders.append({
                'zona': zona,
                'subzona': subzona,
                'tech': tech,
                'status': status,
                'type': order_type,
                'order_num': order_num
            })

    # Group by Zone/Subzone
    tech_orders.sort(key=lambda x: (x['zona'], x['subzona'], x['tech']))
    
    # Generate Output
    print(f"Saving report to {OUTPUT_FILE}...")
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Estado Técnicos por Zona"
    
    # Headers
    headers_out = ['Zona', 'Subzona', 'Técnico asignado', 'Estado Actual', 'Tipo de Orden', 'ID Orden']
    out_ws.append(headers_out)
    
    for item in tech_orders:
        out_ws.append([
            item['zona'],
            item['subzona'],
            item['tech'],
            item['status'],
            item['type'],
            item['order_num']
        ])
        
    out_wb.save(OUTPUT_FILE)
    print("Done.")

if __name__ == "__main__":
    generate_report()
