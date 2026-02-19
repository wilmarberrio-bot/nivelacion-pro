import openpyxl
import os, glob, sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# Find the suggestions file
files = glob.glob("sugerencias_nivelacion*.xlsx")
if files:
    f = sorted(files, key=os.path.getmtime, reverse=True)[0]
    print("=== SUGERENCIAS GENERADAS ===")
    print("File:", f)
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    
    ws = wb["Sugerencias de Movimientos"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print("Headers:", headers)
    for row in ws.iter_rows(min_row=2, values_only=True):
        print("  ", list(row))
    wb.close()

print("\n\n=== ANALISIS DETALLADO DEL ARCHIVO FUENTE ===")
# Now analyze the source data file
files2 = glob.glob("nivelacion*.xlsx")
f2 = sorted(files2, key=os.path.getmtime, reverse=True)[0]
print("Source:", f2)

wb2 = openpyxl.load_workbook(f2, read_only=True, data_only=True)
sheet = wb2.active
headers2 = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

# Build tech data
techs = {}
now_hour = 10.67  # 10:40am approx

for row in sheet.iter_rows(min_row=2, values_only=True):
    tech = str(row[11]).strip() if row[11] else "SIN_ASIGNAR"
    status = str(row[13]).strip() if row[13] else "N/A"
    franja = str(row[9]).strip() if row[9] else "Sin Franja"
    zone = str(row[2]).strip() if row[2] else str(row[8]).strip() if row[8] else "N/A"
    subzone = str(row[1]).strip() if row[1] else "N/A"
    order_id = row[0]
    try:
        lat = float(row[6]) if row[6] else 0
    except:
        lat = 0
    try:
        lon = float(row[7]) if row[7] else 0
    except:
        lon = 0

    if tech not in techs:
        techs[tech] = []
    techs[tech].append({
        "id": order_id,
        "status": status,
        "franja": franja,
        "zone": zone,
        "subzone": subzone,
        "lat": lat,
        "lon": lon
    })

wb2.close()

# Show Yonar detail
print("\n=== YONAR ARRIETA ===")
for t, orders in techs.items():
    if "yonar" in t.lower():
        print(f"  {t}:")
        for o in orders:
            print(f"    ID:{o['id']} Status:{o['status']} Franja:{o['franja']} Zone:{o['zone']} Sub:{o['subzone']}")

# Show all techs with their work status and if they are "about to finish"
STATUS_PROGRESS = {
    'Dispositivos cargados': 4,  # about to finish
    'MAC principal enviada': 3,  # equipment installed
    'Iniciado': 2,               # working inside
    'En sitio': 1,               # just arrived
    'Inbound': 0.5,              # on the way
    'Programado': 0,             # not started
}

print("\n=== TECNICOS POR ZONA CON PROGRESO ===")
for t in sorted(techs.keys()):
    if t == "SIN_ASIGNAR":
        continue
    orders = techs[t]
    total = len(orders)
    zones = set(o['zone'] for o in orders)
    franjas = [o['franja'] for o in orders]
    statuses = [o['status'] for o in orders]
    
    fin = sum(1 for s in statuses if s.lower() in ['finalizado','por auditar'])
    pending = total - fin
    active = [o for o in orders if o['status'].lower() not in ['finalizado','por auditar']]
    prog = [o for o in orders if o['status'].lower() == 'programado']
    
    # Check who's about to finish
    near_done = [o for o in orders if o['status'] in ['Dispositivos cargados', 'MAC principal enviada']]
    
    prog_franjas = [o['franja'] for o in prog]
    
    zone_str = "/".join(zones) if len(zones) <= 2 else f"{len(zones)} zones"
    status_str = ", ".join([f"{s}:{statuses.count(s)}" for s in set(statuses)])
    
    flag = ""
    if near_done:
        flag = " <<< ABOUT TO FINISH"
    if pending <= 2:
        flag += " <<< LOW LOAD"
    
    print(f"  {t}: {total}tot {pending}pend [{status_str}] Zone:{zone_str} Prog:{prog_franjas}{flag}")
