import openpyxl

INPUT_FILE = r"c:\Users\Usuario\OneDrive\Documentos\nivelacion\pre_ruta_2_0_2026-02-17T10_51_45.802611833-05_00.xlsx"

try:
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
    sheet = wb.active
    
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    print(f"Headers: {headers}")
    
    # Find relevant columns
    idx_tech = -1
    idx_status = -1
    for i, h in enumerate(headers):
        if str(h).strip() == 'technenvician':
            idx_tech = i
        if str(h).strip() == 'Estado':
            idx_status = i
            
    if idx_tech != -1 and idx_status != -1:
        empty_tech_count = 0
        programada_empty_tech = 0
        por_asignar_status = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            status = str(row[idx_status]).strip() if row[idx_status] else ""
            tech = str(row[idx_tech]).strip() if row[idx_tech] else ""
            
            if not tech:
                empty_tech_count += 1
                if status == 'Programada':
                    programada_empty_tech += 1
            if status == 'Por asignar':
                por_asignar_status += 1

        print(f"Total rows with EMPTY technician: {empty_tech_count}")
        print(f"Rows with status 'Programada' AND EMPTY technician: {programada_empty_tech}")
        print(f"Rows with status 'Por asignar': {por_asignar_status}")
    else:
        print("Required columns not found.")

except Exception as e:
    print(e)
